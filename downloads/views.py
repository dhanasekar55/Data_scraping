import io
import csv
import re
import zipfile
import pdfplumber
import requests
from django.http import HttpResponse
from django.shortcuts import render
import time
import pandas as pd 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def home(request):
    
    return render(request, 'downloads/home.html')

   #############################################################################################################################################
       
def download_pdfs(request):
    if request.method == 'POST':
        year = request.POST.get('year')
        
        Officer_id_start = request.POST.get('Officer_id_start')
        Officer_id_end = request.POST.get('Officer_id_end')
        start_id =50
        end_id =250

        valid_serial_numbers = []
        count = 0
        
        
        # Check if the user clicked the "Count Certificates" button
        # Check if the user clicked the "Count Certificates" button
        if 'count_certificates' in request.POST:
            base_url = f"https://www.crstn.org/birth_death_tn/CORPBIRTHTAMIL/PDF/death_D-{year}:33"

            for officer_id in range(int(Officer_id_start), int(Officer_id_end) + 1):
                base_urls = f"{base_url}-{officer_id}-"

                for i in range(int(start_id), int(end_id) + 1):
                    pdf_url = f"{base_urls}{i:06d}_old.htm"
                    response = requests.head(pdf_url)  # Only check if the file exists
                    if response.status_code == 200:
                        # Parse the content to extract the necessary data
                        pdf_response = requests.get(pdf_url)
                        soup = BeautifulSoup(pdf_response.content, 'html.parser')
                        td_elements = soup.find_all('td')

                        if len(td_elements) > 1:
                            td_element = td_elements[1]  # 0-based index, so [1] is the second <td>
                            div_element = td_element.find('div')
                            if div_element:
                                spans_in_div = div_element.find_all('span')
                                name = spans_in_div[9].get_text(strip=True) if len(spans_in_div) > 9 else ""
                                gender = spans_in_div[12].get_text(strip=True) if len(spans_in_div) > 12 else ""
                                address = spans_in_div[23].get_text(strip=True) if len(spans_in_div) > 23 else ""
                                # Initialize death date text with a default value
                                death_date_text = "Date not found"

                                # Check index 36 first for the date
                                if len(spans_in_div) > 36:
                                    death_date_text = spans_in_div[36].get_text(strip=True)

                                # If the date from index 36 is not valid, check index 37
                                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', death_date_text)
                                if not date_match and len(spans_in_div) > 37:
                                    death_date_text = spans_in_div[37].get_text(strip=True)

                                # Attempt to extract the date again after checking index 37
                                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', death_date_text)
                                if date_match:
                                    death_date = date_match.group(1)  # Extracted date in the format 09/09/2024
                                else:
                                    death_date = "Date format not recognized"

                                # Append all the details to the valid_serial_numbers list
                                valid_serial_numbers.append({
                                    'year': year,
                                    'officer_id': officer_id,
                                    'certificate_number': f"{i:06d}",
                                    'pdf_url': pdf_url,
                                    'name': name,
                                    'gender':gender,
                                    'death_date': death_date,
                                    'address': address,
                                })
                            count += 1

            # Store valid_serial_numbers in session for later use when downloading
            request.session['valid_serial_numbers'] = valid_serial_numbers

            no_certificates_message = "" if count > 0 else "No certificates found for the provided range."

            return render(request, 'downloads/index.html', {
                'count': count,
                'year': year,
                'serial_numbers': valid_serial_numbers,
                'Officer_id_start': Officer_id_start,
                'Officer_id_end': Officer_id_end,
                'start_id': start_id,
                'end_id': end_id,
                'no_certificates_message': no_certificates_message,
            })

        if 'download_csv' in request.POST:
            # Retrieve valid_serial_numbers from the session
            valid_serial_numbers = request.session.get('valid_serial_numbers', [])

            # Debugging: Check what data is in valid_serial_numbers
            # print(f"Valid serial numbers: {valid_serial_numbers}")
            # print(f"Count of valid serial numbers: {len(valid_serial_numbers)}")

            if not valid_serial_numbers:  # If the list is empty, we can't write to Excel
                return HttpResponse("No valid data to download.", status=400)

            # Create an in-memory output file for the Excel workbook
            output = io.BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Certificates"

            # Write the headers
            headers = ["Year", "Officer ID", "Certificate Number", "Certificate URL", "Name","gender", "Date of Death", "Address"]
            worksheet.append(headers)

            # Write the valid_serial_numbers to Excel
            row_count = 0  # Variable to track the number of rows added
            for index, serial_number in enumerate(valid_serial_numbers, start=1):
                print(f"Adding row {index}: {serial_number}")  # Log each row being added
                worksheet.append([
                    serial_number.get('year', ''),
                    serial_number.get('officer_id', ''),
                    serial_number.get('certificate_number', ''),
                    serial_number.get('pdf_url', ''),
                    serial_number.get('name', ''),
                    serial_number.get('gender', ''),
                    serial_number.get('death_date', ''),
                    serial_number.get('address', ''),
                ])
                row_count += 1  # Increment the row counter

            # Debugging: Check how many rows were added
            # print(f"Rows added to Excel: {row_count}")

            # Adjust the column width for readability
            for col_num, _ in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 20

            # Save the workbook to the output BytesIO
            workbook.save(output)
            output.seek(0)  # Set the file pointer back to the beginning

            # Create the response to download the Excel file
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="certificates.xlsx"'

            return response

    

        
        
        # Handle ZIP download
        if 'download_zip' in request.POST:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            
                base_url = f"https://www.crstn.org/birth_death_tn/CORPBIRTHTAMIL/PDF/death_D-{year}:33"
                
                for officer_id in range(int(Officer_id_start), int(Officer_id_end) + 1):
                    base_urls = f"{base_url}-{officer_id}-"
                    batch_count = 0  # Count the number of certificates in the current batch
                    
                    for i in range(int(start_id), int(end_id) + 1):
                        pdf_url = f"{base_urls}{i:06d}_old.htm"
                        response = requests.get(pdf_url)
                        
                        if response.status_code == 200:
                            zip_file.writestr(f"death_certificate_{year}_{officer_id}_{i:06d}_old.htm", response.content)
                            batch_count += 1
                        
                        # Every 5 certificates, wait for 30 seconds
                        if batch_count == 5:
                            time.sleep(30)  # Delay for 30 seconds
                            batch_count = 0  # Reset the batch count after each batch of 5 certificates
                        else:
                            print(f"Failed to download: {pdf_url} (Status Code: {response.status_code})")
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="certificates.zip"'
            return response

    return render(request, 'downloads/index.html')


   ###################################################################################################################################################
                                                         ##########  2019-2024  ###########

def extract_details_from_certificate(pdf_url):
    """Helper function to extract name, address, and date of death from a PDF URL."""
    response = requests.get(pdf_url)
    
    if response.status_code == 200:
        # Use PDF directly from the response content without saving it
        with pdfplumber.open(io.BytesIO(response.content)) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                full_text += text

        # Extract the name, address, and date of death
        name_line = None
        address_line = None
        death_date_line = None
        gender_line = None

        for line in full_text.split('\n'):
            if "Name" in line and not name_line:  # Extract the first occurrence of "Name"
                name_line = line
            if "Sex" in line and not gender_line:  # We search for the first occurrence of "Sex"
                gender_line = line
            if "Address" in line and not address_line:  # Extract the first occurrence of "Address"
                address_line = line
            if "Date of Death" in line:  # Extract the date of death
                death_date_line = line

        # Extract the name
        name = None
        if name_line:
            name = name_line.split("Name", 1)[1].strip()
        
        gender = None
        if gender_line:
            # Assuming the gender appears after 'Sex'
            gender = gender_line.split("Sex", 1)[1].strip()

        # Extract the address
        address = None
        if address_line:
            # Clean up the address
            address = address_line.replace("Address of the deceased at the time ,", "").strip()
            # Additionally, handle cases where there might be extra commas or whitespace
            address = address.replace(",,", ",").strip()  # Remove double commas
            if address.endswith(','):
                address = address[:-1].strip()  # Remove trailing comma if it exists

        # Extract the date of death
        date_of_death = None
        if death_date_line:
            try:
                date_of_death = death_date_line.split("Date of Death", 1)[1].strip()
            except IndexError:
                date_of_death = death_date_line  # If split fails, return the entire line

        return name, gender, address, date_of_death
    else:
        return None, None, None  # If PDF not accessible



def esign(request):
    if request.method == 'POST':
        year = request.POST.get('year')
        Officer_id_start = request.POST.get('Officer_id_start')
        Officer_id_end = request.POST.get('Officer_id_end')
        start_id = 50
        end_id = 60

        valid_serial_numbers = []
        extracted_data = []
        count = 0
        
        if 'count_certificates' in request.POST:
            base_url = f"https://crstn.org/birth_death_tn/CORPBIRTHTAMIL/esign/signed_death_D-{year}:33"
            
            for officer_id in range(int(Officer_id_start), int(Officer_id_end) + 1):
                base_urls = f"{base_url}-{officer_id}-"
                for i in range(int(start_id), int(end_id) + 1):
                    pdf_url = f"{base_urls}{i:06d}.pdf"
                    response = requests.head(pdf_url)
                    if response.status_code == 200:
                        count += 1
                        
                        name, gender, address, date_of_death = extract_details_from_certificate(pdf_url)
                        
                        extracted_data.append({
                            'serial_number': f"{year}-{officer_id}-{i:06d}",
                            'pdf_url':pdf_url,
                            'name': name or 'Not found',
                            'gender': gender or 'Not found',
                            'date_of_death': date_of_death or 'Not found',
                            'address': address or 'Not found'
                        })
                        
            request.session['extracted_data'] = extracted_data   
            no_certificates_message = "No certificates found for the provided range" if count == 0 else ""

            return render(request, 'downloads/index1.html', {
                'count': count,
                'extracted_data': extracted_data,
                'year': year,
                'Officer_id_start': Officer_id_start,
                'Officer_id_end': Officer_id_end,
                'start_id': start_id,
                'end_id': end_id,
                'no_certificates_message': no_certificates_message,
            })

         # Check if the user clicked the "Download CSV" button
        if 'download_csv' in request.POST:
            # Retrieve the extracted data from the session
            extracted_data = request.session.get('extracted_data', [])

            # Check if there's data to download
            if not extracted_data:
                return HttpResponse("No data available for download.", status=400)

            # Create a DataFrame from the extracted data
            df = pd.DataFrame(extracted_data)

            # Save to Excel
            excel_buffer = io.BytesIO()
            df.to_excel(excel_buffer, index=False, sheet_name='Certificates')
            excel_buffer.seek(0)  # Move the cursor to the start of the BytesIO buffer
            
            # Create the response with the correct headers for an Excel file
            response = HttpResponse(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="certificates.xlsx"'
            return response

    return render(request, 'downloads/index1.html')

###########################################################################################################################################################
##### before 2018 chennai certificates #######
def extract_details(pdf_url):
    # Download the PDF in-memory (no local saving)
    response = requests.get(pdf_url)
    
    if response.status_code == 200:
        # Use PDF directly from the response content without saving it
        with pdfplumber.open(io.BytesIO(response.content)) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:  # Ensure text is not None
                    full_text += text + "\n"  # Add a newline for better separation between pages

        # Extract the name, address, and date of death
        name_line = None
        address_line = None
        death_date_line = None
        gender_line = None  # Initialize gender_line

        # Iterate through the lines to find the name, address, and date of death
        for line in full_text.split('\n'):
            if "Name" in line and not name_line:  # We search for the first occurrence of "Name"
                name_line = line
            if "Sex" in line and not gender_line:  # We search for the first occurrence of "Sex"
                gender_line = line
            if "Address" in line and not address_line:  # Search for any line with "Address"
                address_line = line
            if "Date of Death" in line:  # Search for the date of death line
                death_date_line = line

        # Extract the name
        name = None
        if name_line:
            # Assuming the name appears after 'Name'
            name = name_line.split("Name", 1)[1].strip()
            
        gender = None
        if gender_line:
            # Assuming the gender appears after 'Sex'
            gender = gender_line.split("Sex", 1)[1].strip()

        # Extract the address
        address = None
        if address_line:
            # Clean up the address
            address = address_line.replace("Address of the deceased at the time ,", "").strip()
            # Additionally, handle cases where there might be extra commas or whitespace
            address = address.replace(",,", ",").strip()  # Remove double commas
            if address.endswith(','):
                address = address[:-1].strip()  # Remove trailing comma if it exists

        # Extract the date of death
        date_of_death = None
        if death_date_line:
            try:
                date_of_death = death_date_line.split("Date of Death", 1)[1].strip()
            except IndexError:
                date_of_death = death_date_line  # If the split fails, return the entire line

        # Return name, address, and date of death
        return name, gender, address, date_of_death

    else:
        return None, None, f"Failed to download PDF, status code: {response.status_code}"

def before2018(request):
    no_certificates_message =''
    if request.method == 'POST':
        year = request.POST.get('year')
        # end_year = request.POST.get('end_year')
        zone = request.POST.get('zone')
        division =request.POST.get('division')
        # start_id = request.POST.get('start_id')
        # end_id = request.POST.get('end_id')
        start_id=10
        end_id=20
        

        # Prepare to track the number of valid files
        extracted_datas=[]
        valid_serial_numbers = []
        count = 0
        
        if 'count_certificates' in request.POST:
            base_url = f"https://chennaicorporation.gov.in/online-civic-services/deathCertificate.do?do=DeathCertificate&registrationNumber={year}/{zone}/{division}/"
            
           
            for i in range(int(start_id), int(end_id) + 1):
                    pdf_url = f"{base_url}{i:06d}/0"
                    response = requests.head(pdf_url)
                    if response.status_code == 200:
                        count += 1
                        
                        name, gender, address, date_of_death = extract_details(pdf_url)
                        
                        extracted_datas.append({
                            'serial_number': f"{year}-{zone}-{division}-{i:06d}",
                            'pdf_url':pdf_url,
                            'name': name or 'Not found',
                            'gender': gender or 'Not found',
                            'date_of_death': date_of_death or 'Not found',
                            'address': address or 'Not found'
                        })
                            
            request.session['extracted_datas'] = extracted_datas   
            no_certificates_message = "No certificates found for the provided range" if count == 0 else ""
        if 'download_csv' in request.POST:
            # Retrieve the extracted data from the session
            extracted_datas = request.session.get('extracted_datas', [])

            # Check if there's data to download
            if not extracted_datas:
                return HttpResponse("No data available for download.", status=400)

            # Create a DataFrame from the extracted data
            df = pd.DataFrame(extracted_datas)

            # Save to Excel
            excel_buffer = io.BytesIO()
            df.to_excel(excel_buffer, index=False, sheet_name='Certificates')
            excel_buffer.seek(0)  # Move the cursor to the start of the BytesIO buffer
            
            # Create the response with the correct headers for an Excel file
            response = HttpResponse(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="certificates.xlsx"'
            return response   

        return render(request, 'downloads/index2.html', {
            'count': count,
            'extracted_datas': extracted_datas,
            'year': year,
            'zone': zone,
            'division': division,
            'start_id': start_id,
            'end_id': end_id,
            'no_certificates_message': no_certificates_message,
        })

      
    

    return render(request, 'downloads/index2.html')
